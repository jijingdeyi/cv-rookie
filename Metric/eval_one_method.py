import numpy as np
from PIL import Image
from Metric.metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from typing import Optional

warnings.filterwarnings("ignore")


def _load_or_create_wb(excel_name: str):
    try:
        wb = load_workbook(excel_name)
        # 清理多余的默认Sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])
    except FileNotFoundError:
        wb = Workbook()
        # 新建文件时删除默认Sheet
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)
    return wb


def _get_or_create_ws(wb, worksheet_name: str):
    if worksheet_name in wb.sheetnames:
        ws = wb[worksheet_name]
    else:
        ws = wb.create_sheet(title=worksheet_name)
    return ws


def _first_col_has_values(ws) -> bool:
    # 只要第一列有任意一个非空单元格就视为“已有文件名列”
    for cell in ws.iter_cols(min_col=1, max_col=1, min_row=1, values_only=True):
        if any(v is not None for v in cell):
            return True
    return False


def _next_empty_col_index_0based(ws) -> int:
    """
    找到下一列空白列（0-based）。我们按照 openpyxl 的 max_column 追加即可。
    注意：如果用户中途手动清空了中间某些列，这里仍会在最后一列之后继续追加（更安全）。
    """
    return ws.max_column  # 0-based（因为我们写入时会 +1 变成 1-based 列号）


def _find_col_by_header(ws, header: str, header_row: int = 1) -> Optional[int]:
    target = header.strip().lower()
    hits = []
    for col_idx_1b in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx_1b).value
        if isinstance(val, str) and val.strip().lower() == target:
            hits.append(col_idx_1b - 1)
    if len(hits) > 1:
        print(f"[Warn] 工作表 '{ws.title}' 中存在多个同名方法列 '{header}'，将覆盖第一个匹配列（第 {hits[0]+1} 列）。")
    return hits[0] if hits else None


def write_excel_append(
    excel_name: str,
    worksheet_name: str,
    data: list,
    column_index: int | None = None,
    ensure_filename_first_col: bool = False,
    sanity_name_list: list | None = None,  # 可选：用于做文件名一致性检查
):
    """
    - 若 column_index is None：自动在下一空白列追加（推荐给“方法指标列”用）。
    - 若 ensure_filename_first_col=True：仅当第一列为空时，写入 data 到第一列作为“文件名列”。
      （若第一列已有内容，则跳过，不覆盖）
    - 若 column_index 是显式整数（0-based）：按该列写入。
    """
    wb = _load_or_create_wb(excel_name)
    ws = _get_or_create_ws(wb, worksheet_name)

    # 需要写入文件名列？
    if ensure_filename_first_col:
        if not _first_col_has_values(ws):
            # 第一列为空 -> 写入文件名
            col_letter = get_column_letter(1)  # 第一列
            for i, v in enumerate(data, start=1):
                ws[f"{col_letter}{i}"].value = v
        else:
            # 已有文件名列，若传入 sanity_name_list，可做个简单一致性提示（不抛错）
            if sanity_name_list is not None:
                try:
                    # 取出现有第一列的内容用于比对长度
                    existing_vals = [
                        v[0]
                        for v in ws.iter_rows(
                            min_row=1,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=1,
                            values_only=True,
                        )
                    ]

                    if len(existing_vals) != len(sanity_name_list):
                        print(
                            f"[Warn] 工作表 {worksheet_name} 的文件名列行数({len(existing_vals)})"
                            f"与本次评估数量({len(sanity_name_list)})不一致，请确认是否同一批数据。"
                        )
                except Exception:
                    pass
        # 如果只是确保第一列，直接保存并返回（不去写“方法列”）
        wb.save(excel_name)
        return

    # 计算要写入的目标列
    if column_index is None:
        col0 = _next_empty_col_index_0based(ws)
    else:
        col0 = int(column_index)

    col_letter = get_column_letter(col0 + 1)  # 转 1-based
    for i, v in enumerate(data, start=1):
        ws[f"{col_letter}{i}"].value = v

    # 写完再保险清理
    if (
        "Sheet" in wb.sheetnames
        and worksheet_name != "Sheet"
        and len(wb.sheetnames) > 1
    ):
        wb.remove(wb["Sheet"])
    wb.save(excel_name)


def write_excel_upsert_by_header(
    excel_name: str,
    worksheet_name: str,
    data: list,              # 第一项应为方法名（即你的 Method），后续为各图像指标 + mean + std
    header_row: int = 1,
    clear_trailing: bool = True,  # 若旧列比新列更长，是否清空多余单元格
):
    """
    按第 header_row 行的“表头（方法名）”进行 upsert：
    - 若已存在同名表头：原地覆盖该列数据；
    - 否则：在下一空白列追加。
    """
    wb = _load_or_create_wb(excel_name)
    ws = _get_or_create_ws(wb, worksheet_name)

    if not data:
        wb.save(excel_name)
        return

    header = data[0]  # 你的代码里已在每个指标列顶部插入了 Method
    if not isinstance(header, str) or not header.strip():
        raise ValueError("data 的第一个元素必须是方法名（非空字符串）。")

    # 查找是否已存在同名方法列
    col0 = _find_col_by_header(ws, header, header_row=header_row)
    if col0 is None:
        # 不存在则在下一空列追加
        col0 = _next_empty_col_index_0based(ws)

    # 写入该列（覆盖或追加）
    col_letter = get_column_letter(col0 + 1)  # 转 1-based
    for i, v in enumerate(data, start=1):
        ws[f"{col_letter}{i}"].value = v

    # 如果之前该列行数更长，且需要清理尾部残留，则把多余单元格清空
    if clear_trailing:
        old_max = ws.max_row
        new_len = len(data)
        if old_max > new_len:
            for r in range(new_len + 1, old_max + 1):
                ws.cell(row=r, column=col0 + 1, value=None)

    # 清理默认 Sheet
    if (
        "Sheet" in wb.sheetnames
        and worksheet_name != "Sheet"
        and len(wb.sheetnames) > 1
    ):
        wb.remove(wb["Sheet"])

    wb.save(excel_name)


def evaluation_one(ir_name, vi_name, f_name):
    f_img = Image.open(f_name).convert("L")
    ir_img = Image.open(ir_name).convert("L")
    vi_img = Image.open(vi_name).convert("L")

    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)

    SF = SF_function(f_img_double)
    SD = SD_function(f_img_double)
    AG = AG_function(f_img_double)
    PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)
    MSE = MSE_function(ir_img_double, vi_img_double, f_img_double)
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)
    SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    Nabf = Nabf_function(ir_img_double, vi_img_double, f_img_double)
    SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)
    return EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM


if __name__ == "__main__":
    with_mean = True

    # --- 路径与配置 ---
    Method = "weight"  # 仅用于写入列首和文件名
    ir_dir = os.path.join("/data/ykx/MSRS/test", "ir")
    vi_dir = os.path.join("/data/ykx/MSRS/test", "vi")
    f_dir = os.path.join("/data/ykx/result/ablation", Method)  # 融合结果所在目录
    save_dir = os.path.join("/home/ykx/ReCoNet/result", "metric")
    os.makedirs(save_dir, exist_ok=True)
    metric_save_name = os.path.join(save_dir, f"ablation.xlsx")

    # --- 计算指标 ---
    EN_list, MI_list, SF_list, AG_list, SD_list = [], [], [], [], []
    CC_list, SCD_list, VIF_list = [], [], []
    MSE_list, PSNR_list = [], []
    Qabf_list, Nabf_list = [], []
    SSIM_list, MS_SSIM_list = [], []
    filename_list = [""]  # 第一格空着，对齐“方法名”列

    filelist = natsorted(os.listdir(ir_dir))
    for item in tqdm(filelist, desc=f"Evaluating {Method}"):
        ir_name = os.path.join(ir_dir, item)
        vi_name = os.path.join(vi_dir, item)
        f_name = os.path.join(f_dir, item)
        # 如果融合结果缺失，跳过并给个占位（也可选择 raise）
        if not os.path.exists(f_name):
            # 你也可以选择 continue + 记录缺失文件
            raise FileNotFoundError(f"Fused image not found: {f_name}")

        EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = (
            evaluation_one(ir_name, vi_name, f_name)
        )
        EN_list.append(EN)
        MI_list.append(MI)
        SF_list.append(SF)
        AG_list.append(AG)
        SD_list.append(SD)
        CC_list.append(CC)
        SCD_list.append(SCD)
        VIF_list.append(VIF)
        MSE_list.append(MSE)
        PSNR_list.append(PSNR)
        Qabf_list.append(Qabf)
        Nabf_list.append(Nabf)
        SSIM_list.append(SSIM)
        MS_SSIM_list.append(MS_SSIM)
        filename_list.append(item)

    if with_mean:
        # 用“原始样本”计算统计量，避免被 append 后的统计量污染
        EN_raw, MI_raw, SF_raw, AG_raw, SD_raw = (
            EN_list[:],
            MI_list[:],
            SF_list[:],
            AG_list[:],
            SD_list[:],
        )
        CC_raw, SCD_raw, VIF_raw = CC_list[:], SCD_list[:], VIF_list[:]
        MSE_raw, PSNR_raw = MSE_list[:], PSNR_list[:]
        Qabf_raw, Nabf_raw = Qabf_list[:], Nabf_list[:]
        SSIM_raw, MS_SSIM_raw = SSIM_list[:], MS_SSIM_list[:]

        # 均值
        EN_list.append(np.mean(EN_raw))
        MI_list.append(np.mean(MI_raw))
        SF_list.append(np.mean(SF_raw))
        AG_list.append(np.mean(AG_raw))
        SD_list.append(np.mean(SD_raw))
        CC_list.append(np.mean(CC_raw))
        SCD_list.append(np.mean(SCD_raw))
        VIF_list.append(np.mean(VIF_raw))
        MSE_list.append(np.mean(MSE_raw))
        PSNR_list.append(np.mean(PSNR_raw))
        Qabf_list.append(np.mean(Qabf_raw))
        Nabf_list.append(np.mean(Nabf_raw))
        SSIM_list.append(np.mean(SSIM_raw))
        MS_SSIM_list.append(np.mean(MS_SSIM_raw))
        filename_list.append("mean")

        # 标准差
        EN_list.append(np.std(EN_raw))
        MI_list.append(np.std(MI_raw))
        SF_list.append(np.std(SF_raw))
        AG_list.append(np.std(AG_raw))
        SD_list.append(np.std(SD_raw))
        CC_list.append(np.std(CC_raw))
        SCD_list.append(np.std(SCD_raw))
        VIF_list.append(np.std(VIF_raw))
        MSE_list.append(np.std(MSE_raw))
        PSNR_list.append(np.std(PSNR_raw))
        Qabf_list.append(np.std(Qabf_raw))
        Nabf_list.append(np.std(Nabf_raw))
        SSIM_list.append(np.std(SSIM_raw))
        MS_SSIM_list.append(np.std(MS_SSIM_raw))
        filename_list.append("std")

    # 保留三位小数
    round3 = lambda lst: [round(x, 3) for x in lst]
    EN_list = round3(EN_list)
    MI_list = round3(MI_list)
    SF_list = round3(SF_list)
    AG_list = round3(AG_list)
    SD_list = round3(SD_list)
    CC_list = round3(CC_list)
    SCD_list = round3(SCD_list)
    VIF_list = round3(VIF_list)
    MSE_list = round3(MSE_list)
    PSNR_list = round3(PSNR_list)
    Qabf_list = round3(Qabf_list)
    Nabf_list = round3(Nabf_list)
    SSIM_list = round3(SSIM_list)
    MS_SSIM_list = round3(MS_SSIM_list)

    # 每列顶部插入方法名
    for L in (
        EN_list,
        MI_list,
        SF_list,
        AG_list,
        SD_list,
        CC_list,
        SCD_list,
        VIF_list,
        MSE_list,
        PSNR_list,
        Qabf_list,
        Nabf_list,
        SSIM_list,
        MS_SSIM_list,
    ):
        L.insert(0, Method)

    # 逐表写入
    # —— 先确保每个表的第一列是文件名（若已有则自动跳过，不覆盖）——
    for sheet in [
        "EN",
        "MI",
        "SF",
        "AG",
        "SD",
        "CC",
        "SCD",
        "VIF",
        "MSE",
        "PSNR",
        "Qabf",
        "Nabf",
        "SSIM",
        "MS_SSIM",
    ]:
        write_excel_append(
            metric_save_name,
            sheet,
            data=filename_list,
            ensure_filename_first_col=True,
            sanity_name_list=filename_list,  # 可选：用于简单一致性检查
        )

    # —— 再把本次“方法指标列”自动追加到下一空白列（column_index=None => 追加）——
    sheet_to_list = {
        "EN": EN_list,
        "MI": MI_list,
        "SF": SF_list,
        "AG": AG_list,
        "SD": SD_list,
        "CC": CC_list,
        "SCD": SCD_list,
        "VIF": VIF_list,
        "MSE": MSE_list,
        "PSNR": PSNR_list,
        "Qabf": Qabf_list,
        "Nabf": Nabf_list,
        "SSIM": SSIM_list,
        "MS_SSIM": MS_SSIM_list,
    }
    for sheet, data_list in sheet_to_list.items():
        write_excel_upsert_by_header(
        metric_save_name,
        sheet,
        data=data_list,      # 首元素是 Method，后面是指标 + mean + std
        header_row=1,        # 你的表头就在第1行
        clear_trailing=True  # 如果旧列更长（历史残留），顺手清空尾部
    )
