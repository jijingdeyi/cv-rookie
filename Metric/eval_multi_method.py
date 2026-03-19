import numpy as np
from PIL import Image
from Metric.metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


def get_existing_method_to_col(excel_name: str, worksheet_name: str = "EN") -> dict:
    """
    读取已存在的 Excel，返回 {method_name: column_index}（0-based）。
    约定：第 0 列是文件名；第 1 行（row=1）每个方法列的第一个单元格存 method_name。
    """
    if not os.path.exists(excel_name):
        return {}

    try:
        wb = load_workbook(excel_name, read_only=True, data_only=True)
    except Exception:
        return {}

    if worksheet_name not in wb.sheetnames:
        return {}

    ws = wb[worksheet_name]
    mapping = {}
    # 方法列从第 2 列开始（B 列），A 列是文件名
    col = 2
    while True:
        v = ws.cell(row=1, column=col).value
        if v is None or str(v).strip() == "":
            break
        method = str(v).strip()
        mapping[method] = col - 1  # 转成 0-based column_index
        col += 1
    return mapping


def ensure_filename_column(excel_name: str, filename_list: list, sheet_names: list):
    """
    确保每个 metric sheet 的第 0 列写入 filename_list（A 列）。
    只在 A1 为空/不存在时写，避免覆盖已有表头。
    """
    try:
        wb = load_workbook(excel_name) if os.path.exists(excel_name) else Workbook()
        if wb.active and wb.active.title == "Sheet" and len(wb.sheetnames) == 1:
            # 新建默认空表，先保留，后面会在 write_excel 里清理
            pass
    except FileNotFoundError:
        wb = Workbook()

    for sn in sheet_names:
        ws = wb[sn] if sn in wb.sheetnames else wb.create_sheet(title=sn)
        if ws.cell(row=1, column=1).value in (None, ""):
            col_letter = get_column_letter(1)
            for i, value in enumerate(filename_list, start=1):
                ws[f"{col_letter}{i}"].value = value

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
    wb.save(excel_name)


def write_excel(
    excel_name="metric.xlsx", worksheet_name="VIF", column_index=0, data=None
):
    """将 data 按列写入到指定工作表（0-based 列索引）。会自动移除默认空白 Sheet。"""
    if data is None:
        data = []

    try:
        workbook = load_workbook(excel_name)
        # 如果已有文件里还有默认空白表，且存在其他表，顺手删掉
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            workbook.remove(workbook["Sheet"])
    except FileNotFoundError:
        workbook = Workbook()
        # 新建文件默认会有一个名为 'Sheet' 的空表，直接删掉
        if workbook.active and workbook.active.title == "Sheet":
            workbook.remove(workbook.active)

    # 获取或创建目标工作表
    worksheet = (
        workbook[worksheet_name]
        if worksheet_name in workbook.sheetnames
        else workbook.create_sheet(title=worksheet_name)
    )

    # 在指定列中插入数据（column_index 从 0 开始）
    col_letter = get_column_letter(column_index + 1)
    for i, value in enumerate(data, start=1):
        worksheet[f"{col_letter}{i}"].value = value

    # 为保守起见，如果写完后仍只有一个默认 'Sheet'，也移除
    if (
        "Sheet" in workbook.sheetnames
        and worksheet_name != "Sheet"
        and len(workbook.sheetnames) > 1
    ):
        workbook.remove(workbook["Sheet"])

    workbook.save(excel_name)


def evaluation_one(ir_name, vi_name, f_name):
    f_img = Image.open(f_name).convert("L")
    ir_img = Image.open(ir_name).convert("L")
    vi_img = Image.open(vi_name).convert("L")

    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256, use_nats=True)
    SF = SF_function(f_img_double)
    SD = SD_function(f_img_double)
    AG = AG_function(f_img_double)
    PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)
    MSE = MSE_function(ir_img_double, vi_img_double, f_img_double)
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)
    SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    Nabf = Nabf_function(ir_img_double, vi_img_double, f_img_double)
    SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)
    return EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM


if __name__ == "__main__":
    with_mean = True
    
    ir_dir = os.path.join("/data/ykx/FMB/test", "ir")
    vi_dir = os.path.join("/data/ykx/FMB/test", "vi")
    f_dir = os.path.join('/data/ykx/sota/FMB')
    save_dir = os.path.join("/data/ykx/metric")
    os.makedirs(save_dir, exist_ok=True)

    metric_save_name = os.path.join(save_dir, "FMB.xlsx")
    # filelist = natsorted(os.listdir(ir_dir))
    filelist = os.listdir(ir_dir)

    Method_list = [name for name in os.listdir(f_dir) if os.path.isdir(os.path.join(f_dir, name))]

    metric_sheets = [
        "EN",
        "MI",
        "SF",
        "AG",
        "SD",
        "CC",
        "SCD",
        "VIF",
        "MSE",
        "PSNR",
        "Qabf",
        "Nabf",
        "SSIM",
        "MS_SSIM",
    ]

    # 先根据已有 Excel（默认用 EN sheet）识别已经算过的方法列，避免重算
    existing_method_to_col = get_existing_method_to_col(metric_save_name, worksheet_name="EN")
    existing_methods = set(existing_method_to_col.keys())
    methods_to_run = [m for m in Method_list if m not in existing_methods]

    if len(methods_to_run) == 0:
        print(f"[metric] {metric_save_name} 已包含全部 {len(Method_list)} 个方法，无需重算。")
        raise SystemExit(0)

    # filename 列需要包含 mean/std 行（如果 with_mean=True）
    filename_list = [""] + filelist[:]
    if with_mean:
        filename_list += ["mean", "std"]

    # 确保所有 sheet 的第 0 列（文件名列）存在
    ensure_filename_column(metric_save_name, filename_list, metric_sheets)

    next_col = (max(existing_method_to_col.values()) + 1) if existing_method_to_col else 1

    skipped_methods = []

    for offset, Method in enumerate(methods_to_run):
        sub_f_dir = os.path.join(f_dir, Method)

        # 预检：用第一张图检查融合结果的尺寸是否与源图一致
        first_ir = Image.open(os.path.join(ir_dir, filelist[0]))
        first_f_path = os.path.join(sub_f_dir, filelist[0])
        if not os.path.exists(first_f_path):
            print(f"[skip] {Method}: 融合结果缺失（{first_f_path} 不存在），跳过")
            skipped_methods.append(Method)
            continue
        first_f = Image.open(first_f_path)
        if first_f.size != first_ir.size:
            print(f"[skip] {Method}: 融合图尺寸 {first_f.size} 与源图尺寸 {first_ir.size} 不一致，跳过")
            skipped_methods.append(Method)
            continue

        EN_list, MI_list, SF_list, AG_list, SD_list = [], [], [], [], []
        CC_list, SCD_list, VIF_list = [], [], []
        MSE_list, PSNR_list = [], []
        Qabf_list, Nabf_list = [], []
        SSIM_list, MS_SSIM_list = [], []

        eval_bar = tqdm(filelist, desc=f"{Method}")

        for _, item in enumerate(eval_bar):
            ir_name = os.path.join(ir_dir, item)
            vi_name = os.path.join(vi_dir, item)
            f_name = os.path.join(sub_f_dir, item)
            EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = (
                evaluation_one(ir_name, vi_name, f_name)
            )
            EN_list.append(EN)
            MI_list.append(MI)
            SF_list.append(SF)
            AG_list.append(AG)
            SD_list.append(SD)
            CC_list.append(CC)
            SCD_list.append(SCD)
            VIF_list.append(VIF)
            MSE_list.append(MSE)
            PSNR_list.append(PSNR)
            Qabf_list.append(Qabf)
            Nabf_list.append(Nabf)
            SSIM_list.append(SSIM)
            MS_SSIM_list.append(MS_SSIM)
            filename_list.append(item)

        if with_mean:
            EN_raw, MI_raw, SF_raw, AG_raw, SD_raw = (
                EN_list[:], MI_list[:], SF_list[:], AG_list[:], SD_list[:],
            )
            CC_raw, SCD_raw, VIF_raw = CC_list[:], SCD_list[:], VIF_list[:]
            MSE_raw, PSNR_raw = MSE_list[:], PSNR_list[:]
            Qabf_raw, Nabf_raw = Qabf_list[:], Nabf_list[:]
            SSIM_raw, MS_SSIM_raw = SSIM_list[:], MS_SSIM_list[:]

            EN_list.append(np.mean(EN_raw))
            MI_list.append(np.mean(MI_raw))
            SF_list.append(np.mean(SF_raw))
            AG_list.append(np.mean(AG_raw))
            SD_list.append(np.mean(SD_raw))
            CC_list.append(np.mean(CC_raw))
            SCD_list.append(np.mean(SCD_raw))
            VIF_list.append(np.mean(VIF_raw))
            MSE_list.append(np.mean(MSE_raw))
            PSNR_list.append(np.mean(PSNR_raw))
            Qabf_list.append(np.mean(Qabf_raw))
            Nabf_list.append(np.mean(Nabf_raw))
            SSIM_list.append(np.mean(SSIM_raw))
            MS_SSIM_list.append(np.mean(MS_SSIM_raw))
            filename_list.append("mean")

            EN_list.append(np.std(EN_raw))
            MI_list.append(np.std(MI_raw))
            SF_list.append(np.std(SF_raw))
            AG_list.append(np.std(AG_raw))
            SD_list.append(np.std(SD_raw))
            CC_list.append(np.std(CC_raw))
            SCD_list.append(np.std(SCD_raw))
            VIF_list.append(np.std(VIF_raw))
            MSE_list.append(np.std(MSE_raw))
            PSNR_list.append(np.std(PSNR_raw))
            Qabf_list.append(np.std(Qabf_raw))
            Nabf_list.append(np.std(Nabf_raw))
            SSIM_list.append(np.std(SSIM_raw))
            MS_SSIM_list.append(np.std(MS_SSIM_raw))

        def round3(lst):
            return [round(x, 3) for x in lst]

        EN_list = round3(EN_list)
        MI_list = round3(MI_list)
        SF_list = round3(SF_list)
        AG_list = round3(AG_list)
        SD_list = round3(SD_list)
        CC_list = round3(CC_list)
        SCD_list = round3(SCD_list)
        VIF_list = round3(VIF_list)
        MSE_list = round3(MSE_list)
        PSNR_list = round3(PSNR_list)
        Qabf_list = round3(Qabf_list)
        Nabf_list = round3(Nabf_list)
        SSIM_list = round3(SSIM_list)
        MS_SSIM_list = round3(MS_SSIM_list)

        EN_list.insert(0, f"{Method}")
        MI_list.insert(0, f"{Method}")
        SF_list.insert(0, f"{Method}")
        AG_list.insert(0, f"{Method}")
        SD_list.insert(0, f"{Method}")
        CC_list.insert(0, f"{Method}")
        SCD_list.insert(0, f"{Method}")
        VIF_list.insert(0, f"{Method}")
        MSE_list.insert(0, f"{Method}")
        PSNR_list.insert(0, f"{Method}")
        Qabf_list.insert(0, f"{Method}")
        Nabf_list.insert(0, f"{Method}")
        SSIM_list.insert(0, f"{Method}")
        MS_SSIM_list.insert(0, f"{Method}")

        col_index = next_col + offset - len(skipped_methods)
        write_excel(metric_save_name, "EN", col_index, EN_list)
        write_excel(metric_save_name, "MI", col_index, MI_list)
        write_excel(metric_save_name, "SF", col_index, SF_list)
        write_excel(metric_save_name, "AG", col_index, AG_list)
        write_excel(metric_save_name, "SD", col_index, SD_list)
        write_excel(metric_save_name, "CC", col_index, CC_list)
        write_excel(metric_save_name, "SCD", col_index, SCD_list)
        write_excel(metric_save_name, "VIF", col_index, VIF_list)
        write_excel(metric_save_name, "MSE", col_index, MSE_list)
        write_excel(metric_save_name, "PSNR", col_index, PSNR_list)
        write_excel(metric_save_name, "Qabf", col_index, Qabf_list)
        write_excel(metric_save_name, "Nabf", col_index, Nabf_list)
        write_excel(metric_save_name, "SSIM", col_index, SSIM_list)
        write_excel(metric_save_name, "MS_SSIM", col_index, MS_SSIM_list)
